"""Geracao do relatorio financeiro em Excel."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from config import PASTA_RELATORIOS


AZUL_ESCURO = "17365D"
AZUL = "1F4E78"
AZUL_CLARO = "D9EAF7"
VERDE = "C6EFCE"
VERDE_TEXTO = "006100"
VERMELHO = "FFC7CE"
VERMELHO_TEXTO = "9C0006"
CINZA = "666666"
BRANCO = "FFFFFF"


def criar_resumo_executivo(dados: pd.DataFrame) -> list[str]:
    """Produz destaques automaticos a partir das variacoes do dia."""
    variacoes = dados.dropna(subset=["Variacao Diaria (%)"])
    if variacoes.empty:
        return ["Nao ha variacoes suficientes para gerar os destaques."]

    melhor = variacoes.loc[variacoes["Variacao Diaria (%)"].idxmax()]
    pior = variacoes.loc[variacoes["Variacao Diaria (%)"].idxmin()]
    positivas = int((variacoes["Variacao Diaria (%)"] > 0).sum())
    negativas = int((variacoes["Variacao Diaria (%)"] < 0).sum())
    estaveis = len(variacoes) - positivas - negativas

    return [
        (
            f"Maior alta: {melhor['Ativo']} "
            f"({melhor['Variacao Diaria (%)']:+.2f}%)."
        ),
        (
            f"Maior queda: {pior['Ativo']} "
            f"({pior['Variacao Diaria (%)']:+.2f}%)."
        ),
        (
            f"Panorama: {positivas} ativo(s) em alta, {negativas} em queda "
            f"e {estaveis} estavel(is)."
        ),
    ]


def _ajustar_larguras(planilha) -> None:
    for coluna in planilha.columns:
        maior = max(
            len(str(celula.value)) if celula.value is not None else 0
            for celula in coluna
        )
        letra = get_column_letter(coluna[0].column)
        planilha.column_dimensions[letra].width = min(maior + 3, 32)


def gerar_relatorio(
    dados: pd.DataFrame,
    avisos: list[str] | None = None,
    pasta_saida: Path | None = None,
) -> Path:
    """Cria um arquivo XLSX formatado e retorna seu caminho."""
    if dados.empty:
        raise ValueError("O DataFrame esta vazio; nao ha dados para o relatorio.")

    gerado_em = datetime.now()
    pasta_saida = pasta_saida or PASTA_RELATORIOS
    pasta_saida.mkdir(parents=True, exist_ok=True)
    caminho = pasta_saida / f"relatorio_financeiro_{gerado_em:%Y%m%d_%H%M%S}.xlsx"

    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Mercado Financeiro"
    planilha.sheet_view.showGridLines = False
    planilha.freeze_panes = "A7"

    ultima_coluna = get_column_letter(len(dados.columns))
    planilha.merge_cells(f"A1:{ultima_coluna}1")
    planilha["A1"] = "RELATORIO FINANCEIRO - COTACOES DE MERCADO"
    planilha["A1"].font = Font(size=16, bold=True, color=BRANCO)
    planilha["A1"].fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    planilha["A1"].alignment = Alignment(horizontal="center", vertical="center")
    planilha.row_dimensions[1].height = 28

    planilha.merge_cells(f"A2:{ultima_coluna}2")
    planilha["A2"] = f"Gerado em: {gerado_em:%d/%m/%Y as %H:%M}"
    planilha["A2"].font = Font(italic=True, color=CINZA)
    planilha["A2"].alignment = Alignment(horizontal="center")

    planilha.merge_cells(f"A4:{ultima_coluna}4")
    planilha["A4"] = "Resumo executivo"
    planilha["A4"].font = Font(bold=True, color=BRANCO)
    planilha["A4"].fill = PatternFill("solid", fgColor=AZUL)

    resumo = " | ".join(criar_resumo_executivo(dados))
    planilha.merge_cells(f"A5:{ultima_coluna}5")
    planilha["A5"] = resumo
    planilha["A5"].alignment = Alignment(wrap_text=True, vertical="center")
    planilha["A5"].fill = PatternFill("solid", fgColor=AZUL_CLARO)
    planilha.row_dimensions[5].height = 45

    linha_cabecalho = 7
    borda = Border(bottom=Side(style="thin", color="A6A6A6"))

    for indice, coluna in enumerate(dados.columns, start=1):
        celula = planilha.cell(row=linha_cabecalho, column=indice, value=coluna)
        celula.font = Font(bold=True, color=BRANCO)
        celula.fill = PatternFill("solid", fgColor=AZUL)
        celula.alignment = Alignment(horizontal="center")

    for linha, registro in enumerate(
        dados.itertuples(index=False, name=None),
        start=linha_cabecalho + 1,
    ):
        for coluna, valor in enumerate(registro, start=1):
            celula = planilha.cell(row=linha, column=coluna, value=valor)
            celula.border = borda
            if coluna in (4, 5, 6, 7):
                celula.number_format = '#,##0.00'
            elif coluna == 8:
                celula.number_format = '#,##0'
            elif coluna == 9:
                celula.number_format = "dd/mm/yyyy hh:mm"

    ultima_linha = linha_cabecalho + len(dados)
    tabela = Table(
        displayName="TabelaCotacoes",
        ref=f"A{linha_cabecalho}:{ultima_coluna}{ultima_linha}",
    )
    tabela.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    planilha.add_table(tabela)

    intervalo_variacao = f"E{linha_cabecalho + 1}:E{ultima_linha}"
    planilha.conditional_formatting.add(
        intervalo_variacao,
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            fill=PatternFill("solid", fgColor=VERDE),
            font=Font(color=VERDE_TEXTO),
        ),
    )
    planilha.conditional_formatting.add(
        intervalo_variacao,
        CellIsRule(
            operator="lessThan",
            formula=["0"],
            fill=PatternFill("solid", fgColor=VERMELHO),
            font=Font(color=VERMELHO_TEXTO),
        ),
    )

    if avisos:
        linha_aviso = ultima_linha + 3
        planilha.cell(linha_aviso, 1, "Avisos de coleta").font = Font(
            bold=True,
            color=VERMELHO_TEXTO,
        )
        for deslocamento, aviso in enumerate(avisos, start=1):
            planilha.cell(linha_aviso + deslocamento, 1, f"- {aviso}")

    planilha.auto_filter.ref = f"A{linha_cabecalho}:{ultima_coluna}{ultima_linha}"
    _ajustar_larguras(planilha)
    workbook.save(caminho)
    return caminho
